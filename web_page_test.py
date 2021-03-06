#Python3
import pytest
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import random
#import pygetwindow as gw
import time
#from pynput.keyboard import Key, Controller
import os
from tkinter import *
from openpyxl import load_workbook
#import pygetwindow as gw

curency = ["USD","EUR","GBP","CAD","AUD", "JPY"]

firstname = []
# exel file
file_path_r = (r'C:\Users\PatrykKołodziej\Documents')
os.chdir(file_path_r)

wb = load_workbook('data.xlsx')
ws = wb.active

try:
    wb = load_workbook('data.xlsx')
    ws = wb.active
except:
    pass

sheet1 = wb['Arkusz1']
sheet2 = wb['Arkusz2']

def check_button():

    file_path = file_path_window.get()
    file_name = file_name_window.get()


    isExist_file = os.path.exists(file_path + "/" + file_name)
    isExist_path = os.path.exists(file_path)
    print(isExist_file) # Add this info to APP
    print(isExist_path) # Add this info to APP

    return

def personal_data(country_list,email,name,surname,street,city,phone,cur):
    driver = webdriver.Chrome(executable_path=r"C:\Program Files\Google\Chrome\Application\chromedriver.exe")
    vars = {}
    driver.get("https://dataedo.com/")
    driver.set_window_size(1550, 838)
    driver.find_element(By.LINK_TEXT, "Pricing").click()

    driver.find_element(By.LINK_TEXT, "Quote / Buy").click()


    for i in range(random.randint(0, 9)):
        driver.find_element(By.CSS_SELECTOR, ".ml-1").click()
    driver.find_element(By.CSS_SELECTOR, ".currency").click()
    dropdown = driver.find_element(By.CSS_SELECTOR, ".currency")
    driver.find_element(By.XPATH, "//option[. = '" + cur + "' ]").click()

    driver.find_element(By.CSS_SELECTOR, "option:nth-child(3)").click()

    driver.find_element(By.CSS_SELECTOR, ".btn-primary").click()


    driver.find_element(By.ID, "firstname").click()
    driver.find_element(By.ID, "firstname").send_keys(name)
    if driver.find_element(By.XPATH, "//span[@data-validation-message-for='firstname']").is_displayed():
        print("Test failed first name {}".format(name))
    else:
        print("Test passed first name {}".format(name))

    driver.find_element(By.ID, "lastname").click()
    driver.find_element(By.ID, "lastname").send_keys(surname)

    if driver.find_element(By.XPATH, "//span[@data-validation-message-for='lastname']").is_displayed():
        print("Test failed last name {}".format(surname))

    else:
        print("Test passed last name {}".format(surname))


    driver.find_element(By.ID, "email").click()
    driver.find_element(By.ID, "email").send_keys(email)
    if driver.find_element(By.XPATH, "//span[@data-validation-message-for='email']").is_displayed():
        print("Test failed email {}".format(email))
    else:
        print("Test passed email {}".format(email))


    driver.find_element(By.ID, "phone_no").click()
    driver.find_element(By.ID, "phone_no").send_keys(phone)


    driver.find_element(By.CSS_SELECTOR, ".clearfix:nth-child(2) > .col-md-6").click()
    driver.find_element(By.ID, "ref_no").click()
    driver.find_element(By.ID, "company").click()
    driver.find_element(By.ID, "company").send_keys("company")
    driver.find_element(By.ID, "companydetails").click()
    driver.find_element(By.ID, "companydetails").click()
    driver.find_element(By.ID, "companydetails").click()
    element = driver.find_element(By.ID, "companydetails")
    actions = ActionChains(driver)
    actions.double_click(element).perform()
    driver.find_element(By.ID, "companydetails").send_keys("copmany")
    driver.find_element(By.NAME, "address").click()
    driver.find_element(By.NAME, "address").send_keys(street)
    driver.find_element(By.ID, "addressdetails").click()
    driver.find_element(By.ID, "addressdetails").send_keys(street)
    driver.find_element(By.CSS_SELECTOR, "#address > .form-group:nth-child(5)").click()
    driver.find_element(By.ID, "city").click()
    driver.find_element(By.ID, "city").send_keys("city")
    driver.find_element(By.ID, "citydetails").click()
    driver.find_element(By.ID, "citydetails").send_keys(city)
    driver.find_element(By.ID, "zip").click()
    driver.find_element(By.ID, "zip").send_keys("26-647")
    driver.find_element(By.ID, "zipdetails").click()
    driver.find_element(By.ID, "zipdetails").send_keys("25-647")
    driver.find_element(By.ID, "vat").click()
    driver.find_element(By.ID, "vat").send_keys("PL9591866268")
    try:
        driver.find_element(By.ID, "country").click()
        driver.find_element(By.XPATH, "//option[@value='" + country_list + "']").click()
        print("Test passed country {}".format(country_list))
    except:
        print("Test failed country {}".format(country_list))

def test():
    for i in range(1,10):
        con = sheet2.cell(row=i, column=1).value
        em = sheet1.cell(row=i, column=1).value
        nam = sheet1.cell(row=i, column=2).value
        wam = sheet1.cell(row=i, column=3).value
        str = sheet1.cell(row=i, column=5).value
        cit = sheet1.cell(row=i, column=6).value
        pho = sheet1.cell(row=i, column=8).value
        cur = curency[random.randint(0, 5)]
        personal_data(con, em, nam, wam, str, cit, pho, cur)

root = Tk()
root.title('dataedo')
#File Path
file_path_window = Entry(root, width=35, borderwidth=5)
file_path_window.pack()
file_path_window.insert(0,r"Skopiuj ścieżkę pliku")

file_name_window = Entry(root, width= 35, borderwidth=5)

file_name_window.pack()
file_name_window.insert(0,"Wpisz nazwę pliku z dataedo.xlsx")

x = 1
y = 10
Button_check_path = Button(root, text="Sprawdź plik", padx= 40, pady=20, command=check_button )
Button_check_star = Button(root, text="Start", padx= 40, pady=20, command=test)
Button_check_path.pack()
Button_check_star.pack()

root.mainloop()
